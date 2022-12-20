import openpyxl    

def split_string(string):
    rows = string.split("\n")
    split = [[]]*len(rows)

    for i in range(0,len(rows)):
        split[i] = rows[i].split(" ")
    
    return split

def get_row_format(row):
    # Normal case for values: [label, value_1, value_2, ..., value_no_cols]
    
    type = []
    to_append = ["ppt", "x", "pct", "m", "mm"]
    special = ["-", "n.a.", "na", "NA", "nm", "n.m.", "NM", "–", "—"]

    for i in range(0, len(row)):
        
        # Test whether cell is numeric:

        try:
            num = row[i].replace(")","").replace("(","").replace(",","")
            num.replace("ppt","").replace("%","")
            float(num)
            type.append("num")
        except ValueError:
            if row[i] in to_append:
                type.append("append_prev")
            elif row[i] in special:
                type.append("special")
            else:
                # Check if the string is a label or a number in the label
                if i == 0:
                    # This is the label, so it can contain numbers
                    type.append("label")
                else:
                    # This is not the label, so it should be treated as a string
                    type.append("str")

    return type

def get_header_format(header):

    currency = ["m", "bn", "000", "k", "Mn", "Bn", "M", "B"]
    date = ["H1", "H2", "Q1", "Q2", "Q3", "Q4"]
    format = []

    for i in range(0, len(header)):
        
        print(header[i])

        if any(x in currency for x in header[i]):
            format.append("curr")
        elif any(x in date for x in header[i]) or ("20" in header[i]):
            format.append("date")
        else:
            format.append("word")
    
    return format

def format_header(header):

    format = get_header_format(header)

    if len(format) != len(header):
        print("Error in code in format_header")
        exit()

    result = []
    current = ""

    for i in range(0, len(header)):

        if format[i] == "word" and current == "":
            current += header[i]
        elif format[i] == "word" and current != "":
            current += " " + header[i]
        else:
            if current == "":
                print(f"{header[i]+format[i]}")
                pass
            else:
                result.append(current)
                current = ""

            if format[i] == "date":
                result.append(header[i])
            elif format[i] == "curr":
                try:
                    result[-1] += " " + header[i]
                except IndexError:
                    pass
            else:
                print(f"{header[i]} has unaccounted for format")

    if current != "":
        result.append(current)

    return result

def format_row(row):
    
    format = get_row_format(row) 

    if len(format) != len(row):
        print("Error in code in format_row")
        exit()

    result = []
    current = ""

    for i in range(0, len(format)):
        if format[i] == "str":
            current += " " + row[i]
        elif format[i] == "num":
            result.append(row[i])
        elif format[i] == "append_prev":
            result[-1] += row[i]
        elif format[i] == "label":
            current += " " + row[i]
        else:
            print(f"Other non-supported type in function, {row[i]}")
            result.append(row[i])

    result.insert(0, current[1:])

    return result

def format_table(string):

    split = split_string(string)

    table = []

    for i in range(0, len(split)):
        if i == 0:
            table.append(format_header(split[i]))
        else:       
            table.append(format_row(split[i]))

    return table

def write_to_excel(transformed_list, filename):
    # Create a new workbook
    workbook = openpyxl.Workbook()
    # Get the active sheet
    sheet = workbook.active
    # Iterate over the transformed_list and write the values to the sheet
    for i, row in enumerate(transformed_list):
        for j, value in enumerate(row):
            # try to write the value as a number if it can be interpreted as such
            if value.isnumeric():
                sheet.cell(row=i+1, column=j+1).value = float(value)
            else:
                sheet.cell(row=i+1, column=j+1).value = value
    # Save the workbook
    workbook.save(filename+".xlsx")

def say_hello():
    return "Hello world!"

string = """Year ended 30 June ($m) 2022 2021 Variance % 
Revenue 17,754 16,871 5.2 
EBITDA 3,057 2,993 2.1 
Depreciation and amortisation (740) (692) (6.9) 
EBIT 2,317 2,301 0.7 
Interest on lease liabilities (113) (116) 2.6 
EBT 2,204 2,185 0.9 
Net property contribution 52 (10) n.m. 
EBT (excluding net property contribution) 2,152 2,195 (2.0) 
EBT margin excluding property (%) 12.1 13.0 
ROC (R12, %) 77.2 82.4 
Total store sales growtha (%) 4.2 12.4 
Store-on-store sales growtha,b (%) 4.8 11.9 
Online penetration (%) 3.0 2.3 
Safety (R12, TRIFR) 11.3 11.3 
Scope 1 and 2 emissions, market-based (ktCO2e) 104.9 110.3"""

print(format_table(string))